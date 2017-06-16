#! Python3

"""
Start with a blank worksheet
load list of computers names from AD, SCCM, or JAMF
search work sheet for computer, if found then mark X in the column of what list the computer name is being compared (AD, SCCM, or JAMF)
if not found then add computer to the computer column then mark the column where the computer came from (AD, SCCM, or JAMF)
"""

# Imports
import os, csv, openpyxl, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Protection, Font


# Function
def pop_list():
    """Searches fo CSV files and create a list with its contents"""
    # Create dict_list
    dict_list = {}
    for csvFilename in os.listdir('.'):
        if not csvFilename.endswith('.csv'):
            continue    # skip non-csv files
        name_var = os.path.splitext(csvFilename)[0]
        # print(name_var)
        with open(csvFilename,'r') as csvFileObj:
            readerObj = csv.reader(csvFileObj)
            var_list = []
            for row in readerObj:
                var_list.append(row)
                dict_list[str(name_var)] = var_list
        csvFileObj.close()
    return dict_list

def check_pc(ws, computer_name, column_var):
    # print('Iterating rows in column 1')
    count = 1
    for row in ws.iter_rows(min_row=2, max_col=1):
        # print('Iterating cells in [{0}]'.format(row))
        for cell in row:
            count += 1
            # print(cell)
            # print('Checking cell [{0}] if [{1}] matches [{2}]'.format(cell.coordinate, cell.value, computer_name))
            if cell.value == computer_name:
                row_num = cell.row
                col_num = cell.column
                # print("Match! [{0}] detected in cell [{1}], writing 'X' in cell [{2}]".format(computer_name,cell.coordinate,ws.cell(row=row_num, column=column_var).coordinate))
                ws.cell(row=row_num, column=column_var).value = 'X'
                return [cell, True]
    return [count, False]

def write_pc(ws,computer_name,cell_list, column_var):
    if not cell_list[1]:
        # print("No match found! Writing [{0}] in cell [A{1}] and 'X' in cell [{2}]".format(computer_name, cell_list[0]+1, ws.cell(row=cell_list[0]+1, column=column_var).coordinate))
        ws.cell(row=cell_list[0]+1, column=1).value = computer_name
        ws.cell(row=cell_list[0]+1, column=column_var).value = 'X'

def list_upper(list):
    """Converts list to upper"""
    upper_list = []
    for i in list:
        for j in i:
            upper_list.append(j.upper())
    return upper_list

def loop_pc(ws, dict):
    for list_obj in dict:
        my_var = str(list_obj)
        list_obj = list_upper(dict[list_obj])
        for computer in list_obj:
            if my_var == 'AD_Computers':
                column_var = 2
            elif my_var == 'JAMF_Computers':
                column_var = 3
            elif my_var == 'SCCM_Computers':
                column_var = 4
            cell_obj = check_pc(ws, computer, column_var)
            write_pc(ws, computer, cell_obj, column_var)
    return True

def create_workbook():
    print("Creating Excel Worksheet")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Managed Computers'
    ws['A1'].value = 'Computers'
    ws['A1'].font = Font(bold=True)
    ws['B1'].value = 'AD'
    ws['B1'].font = Font(bold=True)
    ws['C1'].value = 'JAMF'
    ws['C1'].font = Font(bold=True)
    ws['D1'].value = 'SCCM'
    ws['D1'].font = Font(bold=True)
    ws['E1'].value = 'Notes'
    ws['E1'].font = Font(bold=True)
    wb['Managed Computers'].freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:E1'
    name = save_workbook(wb)
    return name

def save_workbook(wb):
    now = datetime.datetime.now().date()
    name = '{0}_COS_Managed_Computers.xlsx'.format(now)
    print("Saving [{0}] Worksheet".format(name))
    wb.save(name)
    return name


def working_directory():
    """Change Working Directory"""
    os.chdir('{0}\\COS Audit\\Managed_Computers\\'.format(os.getcwd()))
    return os.getcwd()

# Main
working_directory()
file_name = create_workbook()
wb = load_workbook(file_name)
ws = wb.get_sheet_by_name('Managed Computers')
dict_list = pop_list()
loop_pc(ws, dict_list)
save_workbook(wb)