#! Python3
"""
1. Grab headers
2. Map headers to db columns
3. Check if key column has value
    if value - new dict > add key (asset_tag) value (WS0002131) to 

    else pass




Start with a blank worksheet
load list of computers names from AD, SCCM, or JAMF
search work sheet for computer, if found then mark X in the column of what list the computer name is being compared (AD, SCCM, or JAMF)
if not found then add computer to the computer column then mark the column where the computer came from (AD, SCCM, or JAMF)
"""
for item 
    item = {'d_code': 'WS00002131', 'd_po':'P0036687'}
    for k,v in item:
        logic


# Imports
import os, csv, openpyxl, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Protection, Font

def list_upper(list):
    """Converts list to upper"""
    upper_list = []
    for i in list:
        for j in i:
            upper_list.append(j.upper())
    return upper_list

def create_worksheet(file_name):
    """Create the excel workbook and worksheet, returns the file name"""
    wb = load_workbook(file_name)
    if 'Computer Assets' in wb.sheetnames:
        ws = wb.get_sheet_by_name('Computer Assets')
    else:
        ws = wb.create_sheet("Computer Assets")
    ws['A1'].value = 'Computers'
    ws['A1'].font = Font(bold=True)
    ws['B1'].value = 'PC Lifecycle'
    ws['B1'].font = Font(bold=True)
    ws['C1'].value = 'Property Control'
    ws['C1'].font = Font(bold=True)
    ws['D1'].value = 'Not Inventoried'
    ws['D1'].font = Font(bold=True)
    ws['E1'].value = 'ORG Code'
    ws['E1'].font = Font(bold=True)
    ws['F1'].value = 'Department'
    ws['F1'].font = Font(bold=True)
    ws['G1'].value = 'Notes'
    ws['G1'].font = Font(bold=True)
    wb['Managed Computers'].freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:E1'
    name = save_workbook(wb)
    return wb

def save_workbook(wb):
    """ Save the workbook and returns the file name"""
    now = datetime.datetime.now().date()
    name = '{0}_COS_Managed_Computers.xlsx'.format(now)
    wb.save(name)
    return name

def root_directory():
    """Root Directory"""
    return os.getcwd()

def working_directory():
    """Change Working Directory"""
    os.chdir('{0}\\COS Audit\\'.format(os.getcwd()))
    return os.getcwd()

def file_directory():
    """Change Working Directory"""
    os.chdir('{0}\\computer_assets\\'.format(os.getcwd()))
    return os.getcwd()

def auto_size_width(ws):
    """Auto size the column width"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 4.5) * 1
        ws.column_dimensions[column].width = adjusted_width

def pop_list():
    """Searches fo CSV files and create a list with its contents"""
    # Create dict_list
    dict_list = {}
    for csvFilename in os.listdir('.'):
        if not csvFilename.endswith('.csv'):
            continue    # skip non-csv files
        name_var = os.path.splitext(csvFilename)[0]
        # print('Getting list of assets from [{0}]'.format(name_var))
        with open(csvFilename,'r') as csvFileObj:
            readerObj = csv.reader(csvFileObj)
            var_list = []
            hrow = next(readerObj) # 1 get headers 2 mapped columb to db 3 
            
            asset_id = hrow.index('D_CODE')
            type_id = hrow.index('D_TYPE')
            serial_id = hrow.index('D_SERIAL_NUM')
            location_id = hrow.index('D_LOCN_CODE_RESP')
            description_id = hrow.index('D_DESC')
            for row in readerObj:
                if row[type_id] == 'PC' or  row[type_id] == 'LT':
                    var_list.append([row[asset_id]])
                    dict_list[str(name_var)] = var_list
        csvFileObj.close()
    return dict_list

def get_file_name():
    now = datetime.datetime.now().date()
    name = '{0}_COS_Managed_Computers.xlsx'.format(now)
    return name

def check_pc(ws, list, list_name, column_var):
    """Checks the worksheet if the computer is already listed and marks it if it does"""
    new_list_name = split_name(list_name)
    count = 1
    for row in ws.iter_rows(min_row=2, max_col=1):
        # print('Iterating cells in [{0}]'.format(row))
        for cell in row:
            count += 1
            # print('Checking cell [{0}] if [{1}] matches [{2}]'.format(cell.coordinate, cell.value, list))
            if cell.value == list:
                row_num = cell.row
                col_num = cell.column
                # print("Match! [{0}] detected in cell [{1}], writing 'X' in cell [{2}]".format(list,cell.coordinate,ws.cell(row=row_num, column=column_var).coordinate))
                ws.cell(row=row_num, column=column_var).value = 'X'
                ws.cell(row=row_num, column=5).value = int(new_list_name[0])
                ws.cell(row=row_num, column=6).value = new_list_name[1]
                return [cell, True]
    return [count, False]

def split_name(list_name):
    new_name = list_name.split('_',1)
    return new_name

def write_pc(ws,asset_tag,cell_list, list_name, column_var):
    """Adds computers into the workbook and marks them"""
    new_list_name = split_name(list_name)
    if not cell_list[1]:
        # print("No match found! Writing [{0}] in cell [A{1}] and 'X' in cell [{2}]".format(asset_tag, cell_list[0]+1, ws.cell(row=cell_list[0]+1, column=column_var).coordinate))
        ws.cell(row=cell_list[0]+1, column=1).value = asset_tag
        ws.cell(row=cell_list[0]+1, column=column_var).value = 'X'
        ws.cell(row=cell_list[0]+1, column=5).value = int(new_list_name[0])
        ws.cell(row=cell_list[0]+1, column=6).value = new_list_name[1]

def loop_pc(ws, dict):
    """Loop through assets and add them to the worksheet"""
    for list_obj in dict:
        list_name = str(list_obj)
        list_obj = list_upper(dict[list_obj])
        for asset in list_obj:
            cell_obj = check_pc(ws, asset, list_name, 2)
            write_pc(ws, asset, cell_obj, list_name, 2)
    return True

# Main Code
# Change to working directory
root = root_directory()
wd = working_directory()
print('Changing to work directory [{0}\\]'.format(wd))

# Create Worksheet
wb = create_worksheet(get_file_name())
ws = wb.get_sheet_by_name('Computer Assets')

# Get computer assets from reports
fd = file_directory()
dict_list = pop_list()
os.chdir(wd)

# Loop through all computers and add them to the worksheet
loop_pc(ws, dict_list)

# Auto size the column width
auto_size_width(ws)

# Auto size the column width
auto_size_width(ws)

# Save the workbook
print("Saving [{0}] Worksheet at [{1}\\]".format(get_file_name(),os.getcwd()))
save_workbook(wb)