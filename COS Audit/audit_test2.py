#! Python3
import os, csv, openpyxl, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Protection, Font

os.chdir('{0}\\COS Audit\\computer_assets\\'.format(os.getcwd()))

wb = load_workbook('template.xlsx')
ws = wb.active
ws.title = 'Computer Assets'
ws['A1'].value = 'Asset Tag'
ws['A1'].font = Font(bold=True)
ws['B1'].value = 'PC Lifecycle'
ws['B1'].font = Font(bold=True)
ws['C1'].value = 'Property Control'
ws['C1'].font = Font(bold=True)
ws['D1'].value = 'Not Inventoried'
ws['D1'].font = Font(bold=True)
ws['E1'].value = 'ORG Code'
ws['E1'].font = Font(bold=True)
ws['F1'].value = 'Department'
ws['F1'].font = Font(bold=True)
ws['G1'].value = 'Type'
ws['G1'].font = Font(bold=True)
ws['H1'].value = 'Searial Number'
ws['H1'].font = Font(bold=True)
ws['I1'].value = 'Location'
ws['I1'].font = Font(bold=True)
ws['J1'].value = 'Description'
ws['J1'].font = Font(bold=True)
ws['K1'].value = 'Notes'
ws['K1'].font = Font(bold=True)
wb['Computer Assets'].freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:K1'
wb.save('test.xlsx')
ws = wb.get_sheet_by_name('Computer Assets')

department_name = '25700_Zoology'
my_dict = {'D_CODE': ['WS0002093', 'WS0002045', 'WS0002095', 'WS0002096', 'WS0002097', 'WS0002098', 'WS0005456', 'WS0004017', 'WS0005137', 'WS0004877', 'WS0006911', 'WS0007541', 'WS0009162', 'WS0004043', 'WS0010740', 'WS0010731', 'WS0010711', 'WS0000988', 'WS0007554', 'WS0012395', 'WS0012396', 'WS0012528', 'WS0012536'], 'D_TYPE': ['PC', 'LT', 'LT', 'LT', 'LT', 'LT', 'LT', 'LT', 'PC', 'LT', 'PC', 'PC', 'PC', 'PC', 'PC', 'PC', 'LT', 'LT', 'PC', 'PC', 'PC', 'LT', 'LT'], 'D_SERIAL_NUM': ['0012254', '090166', '090166', '090168', '090167', '090165', '45ZJLV1', 'C1MGYFSYD', '13143', '9PBLRK1', 'C600HMRY', '095560', '7HK9MB1', 'UPDATEME', '89056', '096714', 'CNU2914PZV', '088912', '096666', 'g6sy63200s0a', 'g6sy63200s09', 'C02R20SCG8WN', 'c02tm0dbgtdx'], 'D_LOCN_CODE_RESP': ['SL0430', 'D20330', 'D20330', 'D20330', 'D20330', 'D20330', 'SL405M', 'SL0409', 'SL0426', 'SL0408', 'SL405M', 'SL0135', 'ET133B', 'SL402A', 'SL0407', 'SL0616', 'SL0428', 'SL0428', 'SL0616', 'TY0435', 'TY0435', 'TY0406', 'TY0407'], 'D_DESC': ['', '', '', '', '', '', '', '', 'Dev math Computer', '', 'Laptop', 'Desktop', 'Dell Optiplex', 'Lenovo Laptop', 'Star West Desktop', 'NUC5i5RYK', 'Laptop', 'Lenovo ThinkPad', 'NUC', 'Molecular Research Lab', 'Molecular Research Lab', 'MACBOOKPRO 2015', 'MBP 13 late 2016"']}


for i in  my_dict['D_CODE']:
    for row in ws.iter_rows(min_row=2, max_col=1):
        print('Iterating cells in [{0}]'.format(row))
        for cell in row:
            print('Checking cell [{0}] if [{1}] matches [{2}]'.format(cell.coordinate, cell.value, i))
            if cell.value == i:
                row_num = cell.row
                print("Match! [{0}] detected in cell [{1}], writing 'X' in cell [{2}]".format(list,cell.coordinate,ws.cell(row=row_num, column=2).coordinate))
                ws.cell(row=row_num, column=2).value = 'X'
                # ws.cell(row=row_num, column=5).value = int(depart_org[0])
                # ws.cell(row=row_num, column=6).value = depart_org[1]
                # ws.cell(row=row_num, column=7).value = depart_org[1]
                # ws.cell(row=row_num, column=8).value = depart_org[1]
                # ws.cell(row=row_num, column=9).value = depart_org[1]
                # ws.cell(row=row_num, column=10).value = depart_org[1]
                wb.save('test.xlsx')