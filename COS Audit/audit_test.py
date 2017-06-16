#! Python3
import os, csv, openpyxl, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Protection, Font

os.chdir('{0}\\COS Audit\\computer_assets\\'.format(os.getcwd()))

wb = load_workbook('template.xlsx')
ws = wb.active
ws.title = 'Computer Assets'
ws['A1'].value = 'Asset Tag'
ws['A1'].font = Font(bold=True)
ws['B1'].value = 'PC Lifecycle'
ws['B1'].font = Font(bold=True)
ws['C1'].value = 'Property Control'
ws['C1'].font = Font(bold=True)
ws['D1'].value = 'Not Inventoried'
ws['D1'].font = Font(bold=True)
ws['E1'].value = 'ORG Code'
ws['E1'].font = Font(bold=True)
ws['F1'].value = 'Department'
ws['F1'].font = Font(bold=True)
ws['G1'].value = 'Type'
ws['G1'].font = Font(bold=True)
ws['H1'].value = 'Searial Number'
ws['H1'].font = Font(bold=True)
ws['I1'].value = 'Location'
ws['I1'].font = Font(bold=True)
ws['J1'].value = 'Description'
ws['J1'].font = Font(bold=True)
ws['K1'].value = 'Notes'
ws['K1'].font = Font(bold=True)
wb['Computer Assets'].freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:K1'
wb.save('test.xlsx')
ws = wb.get_sheet_by_name('Computer Assets')

def populate_dictionary(location):
    """Searches CSV files and create a dictionary list with its contents"""
    dict_list = {}
    for csvFilename in os.listdir(location):
        if not csvFilename.endswith('.csv'):
            continue
        file_name = os.path.splitext(csvFilename)[0]
        dict_list[file_name] = {}
        # print('Getting list of assets from [{0}]'.format(file_name))
        with open(csvFilename,'r') as csvFileObj:
            readerObj = csv.reader(csvFileObj)

            header_row = next(readerObj)

            asset_name = 'D_CODE'
            type_name = 'D_TYPE'
            serial_name = 'D_SERIAL_NUM'
            location_name = 'D_LOCN_CODE_RESP'
            description_name = 'D_DESC'

            asset_id = header_row.index(asset_name)
            type_id = header_row.index(type_name)
            serial_id = header_row.index(serial_name)
            location_id = header_row.index(location_name)
            description_id = header_row.index(description_name)


            for i in header_row:
                if i == asset_name or i == type_name or i == serial_name or i == location_name or i == description_name:
                    dict_list[file_name][i] = []
                    # print('Getting header [{0}] of assets from [{1}]'.format(i,file_name))

            asset_list = []
            type_list = []
            serial_list = []
            location_list = []
            description_list = []

            for row in readerObj:
                if row[type_id] == 'PC' or  row[type_id] == 'LT':
                    asset_list.append(row[asset_id])
                    type_list.append(row[type_id])
                    serial_list.append(row[serial_id])
                    location_list.append(row[location_id])
                    description_list.append(row[description_id])

            dict_list[file_name][asset_name] = asset_list
            dict_list[file_name][type_name] = type_list
            dict_list[file_name][serial_name] = serial_list
            dict_list[file_name][location_name] = location_list
            dict_list[file_name][description_name] = description_list

        csvFileObj.close()
    return dict_list
asset_dictionary = populate_dictionary('C:\\Users\\tylerhardy\\Developer\\pythonProjects\\COS Audit\\computer_assets')

def iterate_assets(dictionary, department, depart_org):
    # print(department)
    pass
    # for i in dictionary[department]['D_CODE']:
        # print(i)
        # for row in ws.iter_rows(min_row=2, max_col=1):
        #     # print('Iterating cells in [{0}]'.format(row))
        #     for cell in row:
        #         # print('Checking cell [{0}] if [{1}] matches [{2}]'.format(cell.coordinate, cell.value, i))
        #         if cell.value == i:
        #             row_num = cell.row
        #             # print("Match! [{0}] detected in cell [{1}], writing 'X' in cell [{2}]".format(list,cell.coordinate,ws.cell(row=row_num, column=2).coordinate))
        #             ws.cell(row=row_num, column=2).value = 'X'
        #             ws.cell(row=row_num, column=5).value = int(depart_org[0])
        #             ws.cell(row=row_num, column=6).value = depart_org[1]
        #             ws.cell(row=row_num, column=7).value = depart_org[1]
        #             ws.cell(row=row_num, column=8).value = depart_org[1]
        #             ws.cell(row=row_num, column=9).value = depart_org[1]
        #             ws.cell(row=row_num, column=10).value = depart_org[1]
        #             wb.save('test.xlsx')
        #             return [cell, True]

# def check_asset()

def loop_department_assets(dictionary):
    # print(dictionary)
    count = 0
    for department in dictionary:
        count += 1
        if count == 2:
            break
        depart_org = department.split('_',1)
        # print(dictionary[department])
        # print(department)
        # print(depart_org)
        # for i in dictionary[department]:
            # print(dictionary[department][i])
            # print(i)
        iterate_assets(dictionary, department, depart_org)
    return True
checking_results = loop_department_assets(asset_dictionary)
# print(asset_dictionary)

department_list = []
for i in asset_dictionary:
    department_list.append(i)
    # print(i)
    # i = asset_dictionary[i]
    # print(i)
# print(department_list)

for i in department_list:
    print(i)
    for j in asset_dictionary[i]:
        print(asset_dictionary[i])

