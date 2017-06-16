#! Python3

# Imports
import os, csv, openpyxl, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Protection, Font

# root = os.getcwd()
# os.chdir('{0}\\COS Audit\\computer_assets\\'.format(root))
os.chdir('{0}\\COS Audit\\test\\'.format(os.getcwd()))

def get_file_name():
    now = datetime.datetime.now().date()
    name = '{0}_COS_Managed_Computers.xlsx'.format(now)
    return name

def create_worksheet(file_name):
    """Create the excel workbook and worksheet, returns the file name"""
    wb = load_workbook(file_name)
    if 'Computer Assets' in wb.sheetnames:
        ws = wb.get_sheet_by_name('Computer Assets')
    else:
        ws = wb.create_sheet("Computer Assets")
    # ws.title = 'Computer Assets'
    ws['A1'].value = 'Computers'
    ws['A1'].font = Font(bold=True)
    ws['B1'].value = 'PC Lifecycle'
    ws['B1'].font = Font(bold=True)
    ws['C1'].value = 'Property Control'
    ws['C1'].font = Font(bold=True)
    ws['D1'].value = 'Org Code'
    ws['D1'].font = Font(bold=True)
    ws['E1'].value = 'Notes'
    ws['E1'].font = Font(bold=True)
    wb['Managed Computers'].freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:E1'
    name = save_workbook(wb)
    return wb



def pop_list():
    """Searches fo CSV files and create a list with its contents"""
    # Create dict_list
    dict_list = {}
    for csvFilename in os.listdir('.'):
        if not csvFilename.endswith('.csv'):
            continue    # skip non-csv files
        name_var = os.path.splitext(csvFilename)[0]
        print('Getting list of assets from [{0}]'.format(name_var))
        # print(name_var)
        var_list = []
        with open(csvFilename,'r') as csvFileObj:
            readerObj = csv.reader(csvFileObj)
            hrow = next(readerObj)
            asset_id = hrow.index('D_CODE')
            type_id = hrow.index('D_TYPE')
            for row in readerObj:
                if row[type_id] == 'PC' or  row[type_id] == 'LT':
                    var_list.append(row[asset_id])
                    dict_list[str(name_var)] = var_list
        csvFileObj.close()
    return dict_list

def loop_pc(ws, dict):
    """Loop through assets and add them to the worksheet"""
    for list_obj in dict:
        print(dict[list_obj])
        # print(list_obj)
        # print(type(list_obj))
        my_var = str(list_obj)
        # list_obj = list_upper(dict[list_obj])
        # for asset in list_obj:
        # print(asset)
        cell_obj = check_pc(ws, dict[list_obj], 2)
        write_pc(ws, dict[list_obj], cell_obj, 2)
    return True

wb = create_worksheet(get_file_name())
ws = wb.get_sheet_by_name('Computer Assets')

asset_list = pop_list()
# print(asset_list)
loop_pc(ws, asset_list)
